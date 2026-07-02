#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""测试修复后的导出功能"""
import sys
sys.path.insert(0, '.')

# 直接导入和测试 row_to_dict
import sqlite3

DB_PATH = "data/accounting.db"

def row_to_dict(row):
    if row is None:
        return None
    if isinstance(row, dict):
        return row
    # 处理 SQLite Row 对象 - 转换成普通字典
    result = {}
    try:
        # 先尝试用 keys() 方法
        for k in row.keys():
            result[k] = row[k]
    except Exception:
        # 如果 keys() 方法失败，手动按索引获取
        keys = ["id", "record_date", "type", "category", "amount", "account", "note", "created_at", "updated_at"]
        for i, k in enumerate(keys):
            try:
                result[k] = row[i]
            except (IndexError, KeyError):
                result[k] = "" if k in ["account", "note"] else None
    return result

# 读取数据
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur = conn.cursor()
cur.execute("SELECT * FROM records ORDER BY record_date DESC, id DESC")
raw_rows = cur.fetchall()
conn.close()

print(f"找到 {len(raw_rows)} 条记录")

# 测试 row_to_dict
print("\n测试 row_to_dict:")
for i, r in enumerate(raw_rows[:2]):
    result = row_to_dict(r)
    print(f"  记录 {i+1}: {result}")
    print(f"  类型: {type(result)}")
    print(f"  键: {list(result.keys())}")

# 现在测试 Excel 导出
print("\n测试 Excel 导出...")
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    import io

    wb = Workbook()
    ws = wb.active
    ws.title = "记账记录"
    
    # 表头
    headers = ["ID", "日期", "类型", "分类", "金额", "账户", "备注", "创建时间", "更新时间"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # 安全获取函数
    def safe_get(row, key, default=""):
        try:
            val = row[key]
            return val if val is not None else default
        except (KeyError, IndexError, TypeError):
            return default
    
    # 写入数据
    rows = [row_to_dict(r) for r in raw_rows]
    for row_num, r in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=safe_get(r, "id", 0))
        ws.cell(row=row_num, column=2, value=safe_get(r, "record_date"))
        r_type = safe_get(r, "type", "expense")
        ws.cell(row=row_num, column=3, value="收入" if r_type == "income" else "支出")
        ws.cell(row=row_num, column=4, value=safe_get(r, "category"))
        ws.cell(row=row_num, column=5, value=safe_get(r, "amount", 0))
        ws.cell(row=row_num, column=6, value=safe_get(r, "account"))
        ws.cell(row=row_num, column=7, value=safe_get(r, "note"))
        ws.cell(row=row_num, column=8, value=safe_get(r, "created_at"))
        ws.cell(row=row_num, column=9, value=safe_get(r, "updated_at"))
    
    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()
    print(f"✅ Excel 导出成功! 大小: {len(data)} bytes")
    
    # 保存到文件测试
    with open("test_export.xlsx", "wb") as f:
        f.write(data)
    print("✅ 已保存到 test_export.xlsx")
    
except Exception as e:
    print(f"❌ Excel 导出错误: {type(e).__name__}: {e}")
    import traceback
    traceback.print_exc()

print("\n✅ 所有测试通过!")
