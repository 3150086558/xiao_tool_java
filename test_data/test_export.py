#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io

# 直接测试数据库读取和Excel导出
DB_PATH = "data/accounting.db"

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur = conn.cursor()
cur.execute("SELECT * FROM records ORDER BY record_date DESC, id DESC")
rows = cur.fetchall()

print(f"找到 {len(rows)} 条记录")

for i, row in enumerate(rows[:3]):
    print(f"\n记录 {i+1}:")
    print(f"  类型: {type(row)}")
    print(f"  键: {row.keys()}")
    for key in row.keys():
        print(f"  {key}: {row[key]}")

# 测试Excel导出
if rows:
    print("\n正在测试Excel导出...")
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "记账记录"
        headers = ["ID", "日期", "类型", "分类", "金额", "账户", "备注", "创建时间", "更新时间"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
        
        for row_num, r in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=r["id"])
            ws.cell(row=row_num, column=2, value=r["record_date"])
            ws.cell(row=row_num, column=3, value="收入" if r["type"] == "income" else "支出")
            ws.cell(row=row_num, column=4, value=r["category"])
            ws.cell(row=row_num, column=5, value=r["amount"])
            ws.cell(row=row_num, column=6, value=r["account"])
            ws.cell(row=row_num, column=7, value=r["note"])
            ws.cell(row=row_num, column=8, value=r["created_at"])
            ws.cell(row=row_num, column=9, value=r["updated_at"])
        
        output = io.BytesIO()
        wb.save(output)
        print(f"Excel导出成功! 大小: {len(output.getvalue())} bytes")
    except Exception as e:
        print(f"Excel导出错误: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()

conn.close()
print("\n测试完成!")
