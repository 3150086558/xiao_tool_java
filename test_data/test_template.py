#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""测试模板生成"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

try:
    wb = Workbook()
    
    # 第1页：填写表格
    ws1 = wb.active
    ws1.title = "导入数据"
    
    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 必填列标记
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 表头
    headers = ["日期", "类型", "项目", "消费分类", "金额", "账户", "备注", "星期几", "是否取消"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        # 必填列标记
        if header in ["日期", "类型", "项目", "金额"]:
            ws1.cell(row=2, column=col, value="* 必填")
            ws1.cell(row=2, column=col).fill = required_fill
            ws1.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        else:
            ws1.cell(row=2, column=col, value="选填")
            ws1.cell(row=2, column=col).alignment = Alignment(horizontal="center")
    
    # 示例数据
    sample_data = [
        ["2026/2/26", "支出", "早餐，一个包子，一个鸡蛋", "变动支出.饮食", -6.85, "微信", "", "星期四", ""],
        ["2026/2/26", "支出", "地铁", "变动支出.交通", -4, "支付宝", "上班通勤", "星期四", ""],
        ["2026/2/27", "收入", "3月工资", "收入.工资", 15000, "银行卡", "2026年3月工资", "星期五", ""],
        ["2026/3/1", "支出", "房租", "固定支出.房租", -2500, "银行卡", "", "星期日", ""],
    ]
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    ws1.column_dimensions["A"].width = 14  # 日期
    ws1.column_dimensions["B"].width = 10  # 类型
    ws1.column_dimensions["C"].width = 28  # 项目
    ws1.column_dimensions["D"].width = 18  # 消费分类
    ws1.column_dimensions["E"].width = 12  # 金额
    ws1.column_dimensions["F"].width = 12  # 账户
    ws1.column_dimensions["G"].width = 25  # 备注
    ws1.column_dimensions["H"].width = 10  # 星期几
    ws1.column_dimensions["I"].width = 10  # 是否取消
    
    # 冻结首行
    ws1.freeze_panes = "A3"
    
    # 第2页：字段说明
    ws2 = wb.create_sheet(title="字段说明")
    
    # 说明标题
    title_font = Font(bold=True, size=14)
    ws2.cell(row=1, column=1, value="记账导入模板 - 字段说明")
    ws2.cell(row=1, column=1).font = title_font
    ws2.merge_cells("A1:D1")
    
    # 字段说明表头
    field_headers = ["字段名称", "是否必填", "格式要求", "示例"]
    for col, header in enumerate(field_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # 字段说明内容
    field_descriptions = [
        ["日期", "必填", "支持多种日期格式\nYYYY/MM/DD\nYYYY-MM-DD\nYYYY年MM月DD日", "2026/2/26\n2026-02-26\n2026年2月26日"],
        ["类型", "必填", "收入 / 支出", "收入\n支出"],
        ["项目", "必填", "消费项目或收入来源", "早餐\n工资\n地铁"],
        ["消费分类", "选填", "更详细的分类标签", "变动支出.饮食\n收入.工资\n固定支出.房租"],
        ["金额", "必填", "数字格式\n负数自动识别为支出", "6.85\n-4.00\n15000"],
        ["账户", "选填", "支付或收款账户", "微信\n支付宝\n银行卡"],
        ["备注", "选填", "额外说明文字", "和同事聚餐\n3月份工资"],
        ["星期几", "选填", "仅供参考，不影响导入", "星期四\n星期五"],
        ["是否取消", "选填", "仅供参考，不影响导入", ""],
    ]
    
    for row_idx, row_data in enumerate(field_descriptions, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 2 and "必填" in str(value):
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # 调整列宽
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 30
    
    # 调整行高
    for row in range(4, 13):
        ws2.row_dimensions[row].height = 45
    
    # 第3页：常见问题
    ws3 = wb.create_sheet(title="常见问题")
    
    qa_title = Font(bold=True, size=12)
    
    qas = [
        ["Q: 日期格式错误怎么办？", "A: 确保填入标准日期格式，如 2026/2/26 或 2026-02-26"],
        ["Q: 金额可以写负数吗？", "A: 可以，负数会自动识别为支出，金额取绝对值"],
        ["Q: 如何区分收入和支出？", "A: 在「类型」列中填写「收入」或「支出」即可"],
        ["Q: 「消费分类」有什么用？", "A: 用于更精细的记账统计，比如「变动支出.饮食」「收入.工资」"],
        ["Q: 导入后可以修改吗？", "A: 可以，导入后可以在网页上编辑或删除"],
        ["Q: 有哪些常用消费分类示例？", "A: 变动支出.饮食、变动支出.交通、变动支出.购物、固定支出.房租、收入.工资、收入.副业"],
    ]
    
    for row_idx, (q, a) in enumerate(qas, 1):
        cell_q = ws3.cell(row=row_idx * 2 - 1, column=1, value=q)
        cell_q.font = qa_title
        cell_q.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        ws3.cell(row=row_idx * 2, column=1, value=a)
    
    ws3.column_dimensions["A"].width = 80
    for row in range(1, len(qas) * 2 + 1):
        ws3.row_dimensions[row].height = 25
    
    # 保存测试
    output = io.BytesIO()
    wb.save(output)
    data = output.getvalue()
    
    print(f"✅ 模板生成成功！大小: {len(data)} bytes")
    
    # 保存到本地文件
    with open("test_template_output.xlsx", "wb") as f:
        f.write(data)
    print("✅ 已保存到 test_template_output.xlsx")
    
except Exception as e:
    import traceback
    print(f"❌ 错误: {e}")
    traceback.print_exc()
