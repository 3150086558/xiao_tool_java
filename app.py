#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
小肖的自用工具
支持 SQLite / PostgreSQL / MySQL

默认 SQLite：python app.py
PostgreSQL：DB_TYPE=postgres DB_HOST=127.0.0.1 DB_PORT=5432 DB_NAME=accounting DB_USER=postgres DB_PASSWORD=123456 python app.py
MySQL：DB_TYPE=mysql DB_HOST=127.0.0.1 DB_PORT=3306 DB_NAME=accounting DB_USER=root DB_PASSWORD=123456 python app.py
"""
import csv
import io
import json
import os
import re
import sqlite3
import hashlib
import secrets
from datetime import datetime
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

BASE_DIR = Path(__file__).resolve().parent
PUBLIC_DIR = BASE_DIR / "public"
DATA_DIR = BASE_DIR / "data"

DB_TYPE = os.getenv("DB_TYPE", "sqlite").strip().lower()
DB_HOST = os.getenv("DB_HOST", "127.0.0.1")
DB_PORT = int(os.getenv("DB_PORT", "5432" if DB_TYPE == "postgres" else "3306"))
DB_NAME = os.getenv("DB_NAME", "accounting")
DB_USER = os.getenv("DB_USER", "")
DB_PASSWORD = os.getenv("DB_PASSWORD", "")
APP_HOST = os.getenv("APP_HOST", "0.0.0.0")
APP_PORT = int(os.getenv("APP_PORT", "8000"))

# 会话存储
SESSIONS = {}

def get_db_connection():
    if DB_TYPE == "sqlite":
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        return sqlite3.connect(DATA_DIR / "0701_my_project.db")
    elif DB_TYPE == "postgres":
        import psycopg2
        return psycopg2.connect(host=DB_HOST, port=DB_PORT, dbname=DB_NAME, user=DB_USER, password=DB_PASSWORD)
    elif DB_TYPE == "mysql":
        import pymysql
        return pymysql.connect(host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASSWORD, database=DB_NAME, autocommit=True)
    else:
        raise ValueError(f"不支持的数据库类型：{DB_TYPE}")

def exec_sql(conn, sql, params=()):
    if DB_TYPE == "sqlite":
        conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute(sql, params)
    return cur

def fetchone(cur):
    row = cur.fetchone()
    if row is None:
        return None
    if DB_TYPE == "sqlite":
        return dict(row)
    cols = [desc[0] for desc in cur.description]
    return dict(zip(cols, row))

def fetchall(cur):
    rows = cur.fetchall()
    if not rows:
        return []
    if DB_TYPE == "sqlite":
        return [dict(row) for row in rows]
    cols = [desc[0] for desc in cur.description]
    return [dict(zip(cols, row)) for row in rows]

def hash_password(password):
    """密码哈希"""
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

def validate_password(password):
    """验证密码规则：8-20位，包含大小写字母和数字"""
    if len(password) < 8 or len(password) > 20:
        return False, "密码长度必须在8-20位之间"
    if not re.search(r'[a-z]', password):
        return False, "密码必须包含小写字母"
    if not re.search(r'[A-Z]', password):
        return False, "密码必须包含大写字母"
    if not re.search(r'[0-9]', password):
        return False, "密码必须包含数字"
    return True, "密码格式正确"

def init_db():
    with get_db_connection() as conn:
        if DB_TYPE == "sqlite":
            # 用户表
            exec_sql(conn, """
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT NOT NULL UNIQUE,
                    password TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
            """);
            # 记账记录表
            exec_sql(conn, """
                CREATE TABLE IF NOT EXISTS records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    record_date TEXT NOT NULL,
                    type TEXT NOT NULL CHECK(type IN ('income','expense')),
                    category TEXT NOT NULL,
                    sub_category TEXT DEFAULT '',
                    amount REAL NOT NULL CHECK(amount >= 0),
                    account TEXT DEFAULT '',
                    note TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            """);
            # 目录表
            exec_sql(conn, """
                CREATE TABLE IF NOT EXISTS menus (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    parent_id INTEGER DEFAULT 0,
                    name TEXT NOT NULL,
                    icon TEXT DEFAULT '',
                    sort_order INTEGER DEFAULT 0,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            """);
            # 待办事项表
            exec_sql(conn, """
                CREATE TABLE IF NOT EXISTS todos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    title TEXT NOT NULL,
                    completed INTEGER DEFAULT 0,
                    priority INTEGER DEFAULT 0,
                    due_date TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            """);
            # 备忘录表
            exec_sql(conn, """
                CREATE TABLE IF NOT EXISTS notes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    title TEXT NOT NULL,
                    content TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            """);
            exec_sql(conn, "CREATE INDEX IF NOT EXISTS idx_records_user ON records(user_id)");
            exec_sql(conn, "CREATE INDEX IF NOT EXISTS idx_records_date ON records(record_date)");
            exec_sql(conn, "CREATE INDEX IF NOT EXISTS idx_menus_user ON menus(user_id)");
            exec_sql(conn, "CREATE INDEX IF NOT EXISTS idx_todos_user ON todos(user_id)");
            exec_sql(conn, "CREATE INDEX IF NOT EXISTS idx_notes_user ON notes(user_id)");
            # 数据库连接配置表
            exec_sql(conn, """
                CREATE TABLE IF NOT EXISTS db_connections (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    name TEXT NOT NULL,
                    db_type TEXT NOT NULL,
                    host TEXT DEFAULT '',
                    port INTEGER DEFAULT 0,
                    username TEXT DEFAULT '',
                    password TEXT DEFAULT '',
                    database TEXT DEFAULT '',
                    sqlite_path TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY (user_id) REFERENCES users(id)
                )
            """)
            exec_sql(conn, "CREATE INDEX IF NOT EXISTS idx_db_conn_user ON db_connections(user_id)")
            conn.commit();
        elif DB_TYPE == "postgres":
            # PostgreSQL实现（略）
            pass
        elif DB_TYPE == "mysql":
            # MySQL实现（略）
            pass

def init_default_menus(user_id):
    """给新用户初始化默认菜单"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with get_db_connection() as conn:
        # 先插入一级菜单
        parent_menus = [
            (user_id, 0, "财务管理", "💰", 1, now),
            (user_id, 0, "日常工具", "🔧", 2, now),
            (user_id, 0, "系统管理", "⚙️", 3, now),
        ]
        parent_ids = []
        for menu in parent_menus:
            exec_sql(conn, """
                INSERT INTO menus(user_id, parent_id, name, icon, sort_order, created_at)
                VALUES(?, ?, ?, ?, ?, ?)
            """, menu)
            cur = exec_sql(conn, "SELECT last_insert_rowid() as id")
            parent_ids.append(fetchone(cur)["id"])
        
        # 用实际插入的 parent_id 插入二级菜单
        child_menus = [
            (user_id, parent_ids[0], "记账", "📊", 1, now),
            (user_id, parent_ids[0], "统计报表", "📈", 2, now),
            (user_id, parent_ids[1], "待办事项", "✅", 1, now),
            (user_id, parent_ids[1], "备忘录", "📝", 2, now),
            (user_id, parent_ids[1], "数据库查询", "🗄️", 3, now),
            (user_id, parent_ids[2], "用户管理", "👥", 1, now),
        ]
        for menu in child_menus:
            exec_sql(conn, """
                INSERT INTO menus(user_id, parent_id, name, icon, sort_order, created_at)
                VALUES(?, ?, ?, ?, ?, ?)
            """, menu)
        conn.commit()

class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(PUBLIC_DIR), **kwargs)
    
    def log_message(self, format, *args):
        pass

    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def get_session_user(self):
        """获取当前登录用户"""
        cookies = self.headers.get("Cookie", "")
        session_id = None
        for cookie in cookies.split(";"):
            if cookie.strip().startswith("session_id="):
                session_id = cookie.strip().split("=", 1)[1]
                break
        if session_id and session_id in SESSIONS:
            return SESSIONS[session_id]
        return None

    def do_GET(self):
        parsed = urlparse(self.path)
        user = self.get_session_user()
        
        # 未登录时跳转到登录页
        if not user and not parsed.path.startswith("/api/") and parsed.path not in ["/login.html", "/register.html"]:
            self.send_response(302)
            self.send_header("Location", "/login.html")
            self.end_headers()
            return
        
        # API路由
        if parsed.path == "/api/health":
            return self.send_json({"ok": True, "db_type": DB_TYPE})
        if parsed.path == "/api/user":
            if user:
                return self.send_json({"ok": True, "user": {"id": user["id"], "username": user["username"]}})
            return self.send_json({"error": "未登录"}, 401)
        if parsed.path == "/api/menus":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_get_menus()
        if parsed.path == "/api/records":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_records(parsed)
        if parsed.path == "/api/summary":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_summary(parsed)
        if parsed.path == "/api/export.csv":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_export_csv(parsed)
        if parsed.path == "/api/export.xlsx":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_export_excel(parsed)
        if parsed.path == "/api/download-template":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_download_template()
        if parsed.path == "/api/todos":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_get_todos()
        if parsed.path == "/api/notes":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_get_notes()
        if parsed.path == "/api/users":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_get_users()
        if parsed.path == "/api/db-connections":
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            return self.api_get_db_connections(user)
        
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        
        if parsed.path == "/api/register":
            return self.api_register()
        if parsed.path == "/api/login":
            return self.api_login()
        if parsed.path == "/api/logout":
            return self.api_logout()
        if parsed.path == "/api/change-password":
            return self.api_change_password()
        
        # 需要登录的API
        user = self.get_session_user()
        if not user:
            return self.send_json({"error": "未登录"}, 401)
        
        if parsed.path == "/api/records":
            return self.api_create_record(user)
        if parsed.path == "/api/import":
            return self.api_import_excel(user)
        if parsed.path == "/api/todos":
            return self.api_create_todo(user)
        if parsed.path == "/api/notes":
            return self.api_create_note(user)
        if parsed.path == "/api/db-query":
            return self.api_db_query(user)
        if parsed.path == "/api/db-connections":
            return self.api_create_db_connection(user)

        return self.send_json({"error": "接口不存在"}, 404)

    def do_PUT(self):
        user = self.get_session_user()
        if not user:
            return self.send_json({"error": "未登录"}, 401)
        
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/records/"):
            return self.api_update_record(user, parsed.path.rsplit("/", 1)[-1])
        if parsed.path.startswith("/api/todos/"):
            return self.api_update_todo(user, parsed.path.rsplit("/", 1)[-1])
        if parsed.path.startswith("/api/notes/"):
            return self.api_update_note(user, parsed.path.rsplit("/", 1)[-1])
        if parsed.path.startswith("/api/users/"):
            action = parsed.path.split("/")[-1]
            return self.api_update_user(user, action)
        if parsed.path.startswith("/api/db-connections/"):
            return self.api_update_db_connection(user, parsed.path.rsplit("/", 1)[-1])
        
        return self.send_json({"error": "接口不存在"}, 404)

    def do_DELETE(self):
        user = self.get_session_user()
        if not user:
            return self.send_json({"error": "未登录"}, 401)
        
        parsed = urlparse(self.path)
        if parsed.path == "/api/clear-all":
            return self.api_clear_all(user)
        if parsed.path.startswith("/api/records/"):
            return self.api_delete_record(user, parsed.path.rsplit("/", 1)[-1])
        if parsed.path.startswith("/api/todos/"):
            return self.api_delete_todo(user, parsed.path.rsplit("/", 1)[-1])
        if parsed.path.startswith("/api/notes/"):
            return self.api_delete_note(user, parsed.path.rsplit("/", 1)[-1])
        if parsed.path.startswith("/api/users/"):
            return self.api_delete_user(user, parsed.path.rsplit("/", 1)[-1])
        if parsed.path.startswith("/api/db-connections/"):
            return self.api_delete_db_connection(user, parsed.path.rsplit("/", 1)[-1])
        
        return self.send_json({"error": "接口不存在"}, 404)

    def read_json(self):
        content_length = int(self.headers.get("Content-Length", 0))
        return json.loads(self.rfile.read(content_length).decode("utf-8"))

    def api_register(self):
        """注册"""
        try:
            data = self.read_json()
            username = str(data.get("username", "")).strip()
            password = str(data.get("password", "")).strip()
            confirm_password = str(data.get("confirm_password", "")).strip()
            
            if not username or len(username) < 3 or len(username) > 20:
                return self.send_json({"error": "用户名长度必须在3-20位之间"}, 400)
            if not re.match(r'^[a-zA-Z0-9_]+$', username):
                return self.send_json({"error": "用户名只能包含字母、数字和下划线"}, 400)
            if password != confirm_password:
                return self.send_json({"error": "两次密码输入不一致"}, 400)
            
            valid, msg = validate_password(password)
            if not valid:
                return self.send_json({"error": msg}, 400)
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            with get_db_connection() as conn:
                # 检查用户名是否存在
                cur = exec_sql(conn, "SELECT id FROM users WHERE username = ?", (username,))
                if fetchone(cur):
                    return self.send_json({"error": "用户名已存在"}, 400)
                
                # 创建用户
                exec_sql(conn, """
                    INSERT INTO users(username, password, created_at)
                    VALUES(?, ?, ?)
                """, (username, hash_password(password), now))
                conn.commit()
                
                # 获取用户ID
                cur = exec_sql(conn, "SELECT id FROM users WHERE username = ?", (username,))
                user = fetchone(cur)
                
                # 初始化默认菜单
                init_default_menus(user["id"])
            
            return self.send_json({"ok": True, "message": "注册成功"})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return self.send_json({"error": str(e)}, 400)

    def api_login(self):
        """登录"""
        try:
            data = self.read_json()
            username = str(data.get("username", "")).strip()
            password = str(data.get("password", "")).strip()
            
            with get_db_connection() as conn:
                cur = exec_sql(conn, "SELECT id, username, password FROM users WHERE username = ?", (username,))
                user = fetchone(cur)
            
            if not user or user["password"] != hash_password(password):
                return self.send_json({"error": "用户名或密码错误"}, 400)
            
            # 创建会话
            session_id = secrets.token_hex(32)
            SESSIONS[session_id] = {"id": user["id"], "username": user["username"]}
            
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Set-Cookie", f"session_id={session_id}; Path=/; HttpOnly; Max-Age=604800")
            body = json.dumps({"ok": True, "user": {"id": user["id"], "username": user["username"]}}, ensure_ascii=False).encode("utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return self.send_json({"error": str(e)}, 400)

    def api_logout(self):
        """登出"""
        cookies = self.headers.get("Cookie", "")
        session_id = None
        for cookie in cookies.split(";"):
            if cookie.strip().startswith("session_id="):
                session_id = cookie.strip().split("=", 1)[1]
                break
        if session_id in SESSIONS:
            del SESSIONS[session_id]
        return self.send_json({"ok": True})

    def api_change_password(self):
        """修改密码"""
        try:
            user = self.get_session_user()
            if not user:
                return self.send_json({"error": "未登录"}, 401)
            
            data = self.read_json()
            old_password = str(data.get("old_password", "")).strip()
            new_password = str(data.get("new_password", "")).strip()
            confirm_password = str(data.get("confirm_password", "")).strip()
            
            if not old_password or not new_password or not confirm_password:
                return self.send_json({"error": "所有字段都不能为空"}, 400)
            
            if new_password != confirm_password:
                return self.send_json({"error": "两次输入的新密码不一致"}, 400)
            
            # 验证新密码格式
            valid, msg = validate_password(new_password)
            if not valid:
                return self.send_json({"error": msg}, 400)
            
            # 验证旧密码
            with get_db_connection() as conn:
                cur = exec_sql(conn, "SELECT password FROM users WHERE id = ?", (user["id"],))
                db_user = fetchone(cur)
                
                if db_user["password"] != hash_password(old_password):
                    return self.send_json({"error": "原密码错误"}, 400)
                
                # 更新密码
                exec_sql(conn, "UPDATE users SET password = ? WHERE id = ?", 
                         (hash_password(new_password), user["id"]))
                conn.commit()
            
            return self.send_json({"ok": True, "message": "密码修改成功"})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return self.send_json({"error": str(e)}, 400)

    def api_get_menus(self):
        """获取菜单列表"""
        user = self.get_session_user()
        with get_db_connection() as conn:
            cur = exec_sql(conn, """
                SELECT * FROM menus WHERE user_id = ? ORDER BY sort_order, id
            """, (user["id"],))
            all_menus = fetchall(cur)
        
        # 构建树形结构
        def build_tree(parent_id=0):
            children = [m for m in all_menus if m["parent_id"] == parent_id]
            for menu in children:
                menu["children"] = build_tree(menu["id"])
            return children
        
        tree = build_tree(0)
        return self.send_json({"menus": tree})

    def parse_filters(self, parsed):
        q = parse_qs(parsed.query)
        return {
            "month": (q.get("month", [""])[0] or "").strip(),
            "type": (q.get("type", [""])[0] or "").strip(),
            "keyword": (q.get("keyword", [""])[0] or "").strip(),
        }

    def build_where(self, filters, user_id):
        where, params = ["user_id = ?"], [user_id]
        if filters["month"]:
            if DB_TYPE == "postgres":
                where.append("to_char(record_date, 'YYYY-MM') = ?")
            elif DB_TYPE == "mysql":
                where.append("date_format(record_date, '%Y-%m') = ?")
            else:
                where.append("strftime('%Y-%m', record_date) = ?")
            params.append(filters["month"])
        if filters["type"]:
            where.append("type = ?")
            params.append(filters["type"])
        if filters["keyword"]:
            where.append("(category LIKE ? OR account LIKE ? OR note LIKE ?)")
            kw = f"%{filters['keyword']}%"
            params.extend([kw, kw, kw])
        return " AND ".join(where), params

    def api_records(self, parsed):
        filters = self.parse_filters(parsed)
        user = self.get_session_user()
        where, params = self.build_where(filters, user["id"])
        with get_db_connection() as conn:
            cur = exec_sql(conn, f"""
                SELECT * FROM records WHERE {where} ORDER BY record_date DESC, id DESC LIMIT 1000
            """, params)
            rows = fetchall(cur)
        self.send_json({"records": rows})

    def api_summary(self, parsed):
        filters = self.parse_filters(parsed)
        user = self.get_session_user()
        where, params = self.build_where(filters, user["id"])
        with get_db_connection() as conn:
            cur = exec_sql(conn, f"""
                SELECT type, COALESCE(SUM(amount), 0) AS total FROM records WHERE {where} GROUP BY type
            """, params)
            summary_rows = fetchall(cur)
            income = sum(r["total"] for r in summary_rows if r["type"] == "income")
            expense = sum(r["total"] for r in summary_rows if r["type"] == "expense")
            
            # 分类汇总
            cur = exec_sql(conn, f"""
                SELECT type, category, COALESCE(SUM(amount), 0) AS amount FROM records WHERE {where}
                GROUP BY type, category ORDER BY amount DESC
            """, params)
            category_rows = fetchall(cur)
        self.send_json({"income": income, "expense": expense, "balance": income - expense, "categories": category_rows})

    def validate_payload(self, data):
        record_date = str(data.get("record_date", "")).strip()
        record_type = str(data.get("type", "")).strip()
        category = str(data.get("category", "")).strip()
        sub_category = str(data.get("sub_category", "")).strip()
        account = str(data.get("account", "")).strip()
        note = str(data.get("note", "")).strip()
        try:
            # 支持多种日期格式
            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
                try:
                    parsed_date = datetime.strptime(record_date, fmt)
                    record_date = parsed_date.strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
            else:
                raise ValueError("日期格式应为 YYYY-MM-DD")
        except ValueError:
            raise ValueError("日期格式应为 YYYY-MM-DD")
        # 支持中文类型
        type_map = {"收入": "income", "支出": "expense", "income": "income", "expense": "expense"}
        if record_type not in type_map:
            raise ValueError("类型只能是收入或支出")
        record_type = type_map[record_type]
        if not category:
            raise ValueError("项目不能为空")
        try:
            amount = round(float(data.get("amount", 0)), 2)
        except Exception:
            raise ValueError("金额必须是数字")
        # 支持负数自动识别类型
        if amount < 0:
            amount = abs(amount)
            if record_type == "income":
                record_type = "expense"
        return record_date, record_type, category, sub_category, amount, account, note

    def api_create_record(self, user):
        try:
            record_date, record_type, category, sub_category, amount, account, note = self.validate_payload(self.read_json())
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                cur = exec_sql(conn, """
                    INSERT INTO records(user_id, record_date, type, category, sub_category, amount, account, note, created_at, updated_at)
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (user["id"], record_date, record_type, category, sub_category, amount, account, note, now, now))
                rid = cur.lastrowid
                conn.commit()
                cur = exec_sql(conn, "SELECT * FROM records WHERE id = ?", (rid,))
                row = fetchone(cur)
            self.send_json({"record": row}, 201)
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_update_record(self, user, record_id):
        try:
            rid = int(record_id)
            record_date, record_type, category, sub_category, amount, account, note = self.validate_payload(self.read_json())
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                cur = exec_sql(conn, """
                    UPDATE records SET record_date=?, type=?, category=?, sub_category=?, amount=?, account=?, note=?, updated_at=?
                    WHERE id=? AND user_id=?
                """, (record_date, record_type, category, sub_category, amount, account, note, now, rid, user["id"]))
                if cur.rowcount == 0:
                    return self.send_json({"error": "记录不存在"}, 404)
                conn.commit()
                cur = exec_sql(conn, "SELECT * FROM records WHERE id=?", (rid,))
                row = fetchone(cur)
            self.send_json({"record": row})
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_delete_record(self, user, record_id):
        try:
            rid = int(record_id)
            with get_db_connection() as conn:
                cur = exec_sql(conn, "DELETE FROM records WHERE id=? AND user_id=?", (rid, user["id"]))
                if cur.rowcount == 0:
                    return self.send_json({"error": "记录不存在"}, 404)
                conn.commit()
            self.send_json({"ok": True})
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_clear_all(self, user):
        """清空当前用户所有数据"""
        try:
            with get_db_connection() as conn:
                cur = exec_sql(conn, "DELETE FROM records WHERE user_id=?", (user["id"],))
                deleted_count = cur.rowcount
                conn.commit()
            self.send_json({"ok": True, "deleted_count": deleted_count})
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    # ========== 待办事项 API ==========
    def api_get_todos(self):
        user = self.get_session_user()
        with get_db_connection() as conn:
            cur = exec_sql(conn, """
                SELECT * FROM todos WHERE user_id = ? ORDER BY completed ASC, priority DESC, created_at DESC
            """, (user["id"],))
            rows = fetchall(cur)
        self.send_json({"todos": rows})

    def api_create_todo(self, user):
        try:
            data = self.read_json()
            title = str(data.get("title", "")).strip()
            priority = int(data.get("priority", 0))
            due_date = str(data.get("due_date", "")) or None
            
            if not title:
                return self.send_json({"error": "标题不能为空"}, 400)
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                cur = exec_sql(conn, """
                    INSERT INTO todos(user_id, title, priority, due_date, created_at, updated_at)
                    VALUES(?, ?, ?, ?, ?, ?)
                """, (user["id"], title, priority, due_date, now, now))
                rid = cur.lastrowid
                conn.commit()
                cur = exec_sql(conn, "SELECT * FROM todos WHERE id = ?", (rid,))
                row = fetchone(cur)
            self.send_json({"todo": row}, 201)
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_update_todo(self, user, todo_id):
        try:
            rid = int(todo_id)
            data = self.read_json()
            title = str(data.get("title", "")).strip()
            completed = int(data.get("completed", 0))
            priority = int(data.get("priority", 0))
            due_date = str(data.get("due_date", "")) or None
            
            if not title:
                return self.send_json({"error": "标题不能为空"}, 400)
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                cur = exec_sql(conn, """
                    UPDATE todos SET title=?, completed=?, priority=?, due_date=?, updated_at=?
                    WHERE id=? AND user_id=?
                """, (title, completed, priority, due_date, now, rid, user["id"]))
                if cur.rowcount == 0:
                    return self.send_json({"error": "待办事项不存在"}, 404)
                conn.commit()
                cur = exec_sql(conn, "SELECT * FROM todos WHERE id=?", (rid,))
                row = fetchone(cur)
            self.send_json({"todo": row})
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_delete_todo(self, user, todo_id):
        try:
            rid = int(todo_id)
            with get_db_connection() as conn:
                cur = exec_sql(conn, "DELETE FROM todos WHERE id=? AND user_id=?", (rid, user["id"]))
                if cur.rowcount == 0:
                    return self.send_json({"error": "待办事项不存在"}, 404)
                conn.commit()
            self.send_json({"ok": True})
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    # ========== 备忘录 API ==========
    def api_get_notes(self):
        user = self.get_session_user()
        with get_db_connection() as conn:
            cur = exec_sql(conn, """
                SELECT * FROM notes WHERE user_id = ? ORDER BY updated_at DESC
            """, (user["id"],))
            rows = fetchall(cur)
        self.send_json({"notes": rows})

    def api_create_note(self, user):
        try:
            data = self.read_json()
            title = str(data.get("title", "")).strip()
            content = str(data.get("content", "")).strip()
            
            if not title:
                return self.send_json({"error": "标题不能为空"}, 400)
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                cur = exec_sql(conn, """
                    INSERT INTO notes(user_id, title, content, created_at, updated_at)
                    VALUES(?, ?, ?, ?, ?)
                """, (user["id"], title, content, now, now))
                rid = cur.lastrowid
                conn.commit()
                cur = exec_sql(conn, "SELECT * FROM notes WHERE id = ?", (rid,))
                row = fetchone(cur)
            self.send_json({"note": row}, 201)
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_update_note(self, user, note_id):
        try:
            rid = int(note_id)
            data = self.read_json()
            title = str(data.get("title", "")).strip()
            content = str(data.get("content", "")).strip()
            
            if not title:
                return self.send_json({"error": "标题不能为空"}, 400)
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                cur = exec_sql(conn, """
                    UPDATE notes SET title=?, content=?, updated_at=?
                    WHERE id=? AND user_id=?
                """, (title, content, now, rid, user["id"]))
                if cur.rowcount == 0:
                    return self.send_json({"error": "备忘录不存在"}, 404)
                conn.commit()
                cur = exec_sql(conn, "SELECT * FROM notes WHERE id=?", (rid,))
                row = fetchone(cur)
            self.send_json({"note": row})
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_delete_note(self, user, note_id):
        try:
            rid = int(note_id)
            with get_db_connection() as conn:
                cur = exec_sql(conn, "DELETE FROM notes WHERE id=? AND user_id=?", (rid, user["id"]))
                if cur.rowcount == 0:
                    return self.send_json({"error": "备忘录不存在"}, 404)
                conn.commit()
            self.send_json({"ok": True})
        except Exception as e:
            self.send_json({"error": str(e)}, 400)

    def api_export_csv(self, parsed):
        filters = self.parse_filters(parsed)
        user = self.get_session_user()
        where, params = self.build_where(filters, user["id"])
        with get_db_connection() as conn:
            cur = exec_sql(conn, f"SELECT * FROM records WHERE {where} ORDER BY record_date DESC, id DESC", params)
            rows = fetchall(cur)
        out = io.StringIO()
        writer = csv.writer(out)
        writer.writerow(["ID", "日期", "类型", "项目", "金额", "消费分类", "账户", "备注", "创建时间", "更新时间"])
        for r in rows:
            writer.writerow([r["id"], r["record_date"], "收入" if r["type"] == "income" else "支出", r["category"], r["amount"], r.get("sub_category", ""), r.get("account", ""), r.get("note", ""), r["created_at"], r["updated_at"]])
        body = "\ufeff" + out.getvalue()
        data = body.encode("utf-8")
        self.send_response(200)
        self.send_header("Content-Type", "text/csv; charset=utf-8")
        self.send_header("Content-Disposition", "attachment; filename=accounting-records.csv")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def api_export_excel(self, parsed):
        if not EXCEL_AVAILABLE:
            return self.send_json({"error": "Excel 导出功能不可用，请先安装 openpyxl 库"}, 500)
        filters = self.parse_filters(parsed)
        user = self.get_session_user()
        where, params = self.build_where(filters, user["id"])
        with get_db_connection() as conn:
            cur = exec_sql(conn, f"SELECT * FROM records WHERE {where} ORDER BY record_date DESC, id DESC", params)
            rows = fetchall(cur)
        wb = Workbook()
        ws = wb.active
        ws.title = "记账记录"
        # 表头样式
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")
        # 写入表头
        headers = ["ID", "日期", "类型", "项目", "金额", "消费分类", "账户", "备注", "创建时间", "更新时间"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        # 写入数据
        def safe_get(row, key, default=""):
            try:
                val = row[key]
                return val if val is not None else default
            except (KeyError, IndexError, TypeError):
                return default
        for row_num, r in enumerate(rows, 2):
            ws.cell(row=row_num, column=1, value=safe_get(r, "id", 0))
            ws.cell(row=row_num, column=2, value=safe_get(r, "record_date"))
            ws.cell(row=row_num, column=3, value="收入" if safe_get(r, "type") == "income" else "支出")
            ws.cell(row=row_num, column=4, value=safe_get(r, "category"))
            ws.cell(row=row_num, column=5, value=safe_get(r, "amount", 0))
            ws.cell(row=row_num, column=6, value=safe_get(r, "sub_category"))
            ws.cell(row=row_num, column=7, value=safe_get(r, "account"))
            ws.cell(row=row_num, column=8, value=safe_get(r, "note"))
            ws.cell(row=row_num, column=9, value=safe_get(r, "created_at"))
            ws.cell(row=row_num, column=10, value=safe_get(r, "updated_at"))
        output = io.BytesIO()
        wb.save(output)
        data = output.getvalue()
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=accounting-records.xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def api_download_template(self):
        """下载导入模板"""
        if not EXCEL_AVAILABLE:
            return self.send_json({"error": "Excel 功能不可用，请安装 openpyxl"}, 500)
        from openpyxl.styles import PatternFill
        
        wb = Workbook()
        
        # ====== 第1页：填写表格 ======
        ws1 = wb.active
        ws1.title = "导入数据"
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 必填列标记
        required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # 表头
        headers = ["日期", "类型", "项目", "金额", "消费分类", "账户", "备注", "星期几", "是否取消"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            # 必填列标记
            if header in ["日期", "类型", "项目", "金额"]:
                ws1.cell(row=2, column=col, value="* 必填")
                ws1.cell(row=2, column=col).fill = required_fill
                ws1.cell(row=2, column=col).alignment = Alignment(horizontal="center")
            else:
                ws1.cell(row=2, column=col, value="选填")
                ws1.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        
        # 示例数据
        sample_data = [
            ["2025/06/17", "支出", "午餐", 6.85, "餐饮", "微信", "", "星期四", ""],
            ["2025/06/17", "支出", "地铁", 4, "交通", "支付宝", "上班通勤", "星期四", ""],
            ["2025/06/17", "收入", "工资", 15000, "收入", "银行卡", "2025年6月工资", "星期五", ""],
        ]
        for row_idx, row_data in enumerate(sample_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
        
        # 调整列宽
        ws1.column_dimensions["A"].width = 14  # 日期
        ws1.column_dimensions["B"].width = 10  # 类型
        ws1.column_dimensions["C"].width = 28  # 项目
        ws1.column_dimensions["D"].width = 12  # 金额
        ws1.column_dimensions["E"].width = 18  # 消费分类
        ws1.column_dimensions["F"].width = 12  # 账户
        ws1.column_dimensions["G"].width = 25  # 备注
        
        # 冻结首行
        ws1.freeze_panes = "A3"
        
        # ====== 第2页：字段说明 ======
        ws2 = wb.create_sheet(title="字段说明")
        
        # 说明标题
        title_font = Font(bold=True, size=14)
        ws2.cell(row=1, column=1, value="记账导入模板 - 字段说明")
        ws2.cell(row=1, column=1).font = title_font
        ws2.merge_cells("A1:D1")
        
        # 字段说明表头
        field_headers = ["字段名称", "是否必填", "格式要求", "示例"]
        for col, header in enumerate(field_headers, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 字段说明内容
        field_descriptions = [
            ["日期", "必填", "支持多种日期格式\nYYYY/MM/DD\nYYYY-MM-DD\nYYYY年MM月DD日", "2025/06/17\n2025-06-17\n2025年6月17日"],
            ["类型", "必填", "收入 / 支出", "收入\n支出"],
            ["项目", "必填", "消费项目或收入来源", "午餐\n工资\n地铁"],
            ["金额", "必填", "数字格式\n负数自动识别为支出", "6.85\n-4.00\n15000"],
            ["消费分类", "选填", "更详细的分类标签", "餐饮\n收入\n交通"],
            ["账户", "选填", "支付或收款账户", "微信\n支付宝\n银行卡"],
            ["备注", "选填", "额外说明文字", "和同事聚餐"],
        ]
        
        for row_idx, row_data in enumerate(field_descriptions, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col_idx == 2 and "必填" in str(value):
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # 调整列宽
        ws2.column_dimensions["A"].width = 14
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 35
        ws2.column_dimensions["D"].width = 30
        
        # 调整行高
        for row in range(4, 11):
            ws2.row_dimensions[row].height = 45
        
        output = io.BytesIO()
        wb.save(output)
        data = output.getvalue()
        
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", "attachment; filename=AccountingTemplate.xlsx")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def api_import_excel(self, user):
        """导入Excel文件"""
        if not EXCEL_AVAILABLE:
            return self.send_json({"error": "Excel 功能不可用，请安装 openpyxl"}, 500)
        try:
            # 读取 multipart/form-data
            content_type = self.headers.get("Content-Type", "")
            if "multipart/form-data" not in content_type:
                return self.send_json({"error": "请上传 Excel 文件"}, 400)
            
            # 获取 boundary
            boundary = content_type.split("boundary=")[-1].encode()
            
            # 读取请求体
            content_length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(content_length)
            
            # 简单解析 multipart 数据
            parts = body.split(b"--" + boundary)
            file_data = None
            for part in parts:
                if b"filename=" in part and b"\r\n\r\n" in part:
                    file_data = part.split(b"\r\n\r\n")[1].rsplit(b"\r\n", 1)[0]
                    break
            
            if not file_data:
                return self.send_json({"error": "未找到文件"}, 400)
            
            # 解析 Excel
            from openpyxl import load_workbook
            excel_file = io.BytesIO(file_data)
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # 读取表头，建立字段映射
            headers = {}
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(row=1, column=col).value or "").strip()
                if val:
                    headers[val] = col
            
            # 必须字段
            required = ["日期", "类型", "项目", "金额"]
            for r in required:
                if r not in headers:
                    return self.send_json({"error": f"缺少必须列：{r}"}, 400)
            
            # 读取数据行
            success_count = 0
            errors = []
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            with get_db_connection() as conn:
                for row_num in range(3, ws.max_row + 1):
                    try:
                        # 读取各字段
                        record_date = str(ws.cell(row=row_num, column=headers["日期"]).value or "").strip()
                        record_type = str(ws.cell(row=row_num, column=headers["类型"]).value or "").strip()
                        category = str(ws.cell(row=row_num, column=headers["项目"]).value or "").strip()
                        amount_val = ws.cell(row=row_num, column=headers["金额"]).value
                        
                        # 可选字段
                        sub_category = str(ws.cell(row=row_num, column=headers.get("消费分类", 0)).value or "").strip() if headers.get("消费分类") else ""
                        note = str(ws.cell(row=row_num, column=headers.get("备注", 0)).value or "").strip() if headers.get("备注") else ""
                        account = str(ws.cell(row=row_num, column=headers.get("账户", 0)).value or "").strip() if headers.get("账户") else ""
                        
                        # 跳过空行
                        has_data = (record_date and str(record_date).strip()) or (category and str(category).strip())
                        if not has_data:
                            continue
                        
                        # 解析日期
                        parsed_date = None
                        if isinstance(record_date, datetime):
                            parsed_date = record_date
                        else:
                            for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S"):
                                try:
                                    parsed_date = datetime.strptime(str(record_date), fmt)
                                    break
                                except ValueError:
                                    continue
                        
                        if not parsed_date:
                            errors.append(f"第 {row_num} 行：日期格式无效")
                            continue
                        
                        record_date = parsed_date.strftime("%Y-%m-%d")
                        
                        type_map = {"收入": "income", "支出": "expense"}
                        if record_type not in type_map:
                            errors.append(f"第 {row_num} 行：类型只能是收入或支出")
                            continue
                        record_type = type_map[record_type]
                        
                        try:
                            amount = round(float(amount_val or 0), 2)
                        except (ValueError, TypeError):
                            errors.append(f"第 {row_num} 行：金额格式无效")
                            continue
                        
                        if amount < 0:
                            amount = abs(amount)
                            if record_type == "income":
                                record_type = "expense"
                        
                        if not category:
                            errors.append(f"第 {row_num} 行：项目不能为空")
                            continue
                        
                        exec_sql(conn, """
                            INSERT INTO records(user_id, record_date, type, category, sub_category, amount, account, note, created_at, updated_at)
                            VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (user["id"], record_date, record_type, category, sub_category, amount, account, note, now, now))
                        success_count += 1
                        
                    except Exception as e:
                        errors.append(f"第 {row_num} 行：{str(e)}")
                
                conn.commit()
            
            self.send_json({"success": success_count, "errors": errors})
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return self.send_json({"error": str(e)}, 400)

    def api_get_users(self):
        """获取用户列表（仅admin可用）"""
        user = self.get_session_user()
        if user["username"] != "admin":
            return self.send_json({"error": "仅管理员可访问"}, 403)
        with get_db_connection() as conn:
            cur = exec_sql(conn, """
                SELECT u.id, u.username, u.created_at,
                    COUNT(DISTINCT r.id) as record_count,
                    COUNT(DISTINCT t.id) as todo_count,
                    COUNT(DISTINCT n.id) as note_count
                FROM users u
                LEFT JOIN records r ON r.user_id = u.id
                LEFT JOIN todos t ON t.user_id = u.id
                LEFT JOIN notes n ON n.user_id = u.id
                GROUP BY u.id
                ORDER BY u.id
            """)
            users = fetchall(cur)
        return self.send_json({"users": users})

    def api_delete_user(self, admin_user, target_id):
        """删除用户（仅admin可用，不能删自己）"""
        if admin_user["username"] != "admin":
            return self.send_json({"error": "仅管理员可操作"}, 403)
        target_id = int(target_id)
        if target_id == admin_user["id"]:
            return self.send_json({"error": "不能删除自己"}, 400)
        with get_db_connection() as conn:
            cur = exec_sql(conn, "SELECT id FROM users WHERE id = ?", (target_id,))
            if not fetchone(cur):
                return self.send_json({"error": "用户不存在"}, 404)
            # 删除用户的所有关联数据
            exec_sql(conn, "DELETE FROM notes WHERE user_id = ?", (target_id,))
            exec_sql(conn, "DELETE FROM todos WHERE user_id = ?", (target_id,))
            exec_sql(conn, "DELETE FROM records WHERE user_id = ?", (target_id,))
            exec_sql(conn, "DELETE FROM menus WHERE user_id = ?", (target_id,))
            exec_sql(conn, "DELETE FROM users WHERE id = ?", (target_id,))
            conn.commit()
        return self.send_json({"ok": True, "message": "用户已删除"})

    def api_update_user(self, admin_user, action):
        """管理员更新用户（重置密码）"""
        if admin_user["username"] != "admin":
            return self.send_json({"error": "仅管理员可操作"}, 403)
        try:
            data = self.read_json()
            target_id = data.get("user_id")
            new_password = data.get("new_password", "").strip()

            if not target_id:
                return self.send_json({"error": "缺少用户ID"}, 400)
            if action != "reset-password":
                return self.send_json({"error": "未知操作"}, 404)

            valid, msg = validate_password(new_password)
            if not valid:
                return self.send_json({"error": msg}, 400)

            with get_db_connection() as conn:
                cur = exec_sql(conn, "SELECT id FROM users WHERE id = ?", (target_id,))
                if not fetchone(cur):
                    return self.send_json({"error": "用户不存在"}, 404)
                exec_sql(conn, "UPDATE users SET password = ? WHERE id = ?",
                         (hash_password(new_password), target_id))
                conn.commit()
            return self.send_json({"ok": True, "message": "密码已重置"})
        except Exception as e:
            return self.send_json({"error": str(e)}, 400)

    def _connect_external_db(self, config):
        """连接外部数据库"""
        db_type = config.get("db_type", "postgres").lower()

        if db_type == "sqlite":
            import sqlite3
            sqlite_path = config.get("sqlite_path", "")
            if not sqlite_path:
                raise ValueError("SQLite 数据库文件路径不能为空")
            if not os.path.exists(sqlite_path):
                raise FileNotFoundError(f"SQLite 文件不存在：{sqlite_path}")
            conn = sqlite3.connect(sqlite_path)
            conn.row_factory = sqlite3.Row
            return conn
        elif db_type == "postgres":
            import psycopg2, socket
            host = config.get("host", "127.0.0.1")
            port = int(config.get("port", 5432))
            user = config.get("username", "")
            password = config.get("password", "")
            database = config.get("database", "")
            # 先用 socket 探测端口是否可达
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(3)
                sock.connect((host, port))
                sock.close()
            except Exception:
                raise Exception(f"PostgreSQL 连接失败：无法访问 {host}:{port}，请确认 PostgreSQL 服务已启动且端口正确")
            try:
                return psycopg2.connect(host=host, port=port, dbname=database, user=user, password=password)
            except psycopg2.OperationalError:
                # Windows 下 psycopg2 OperationalError 的 args 和 str() 均为空，无法获取具体原因
                raise Exception(f"PostgreSQL 连接失败：无法连接到 {host}:{port} 上的数据库 \"{database}\"，请检查用户名、密码是否正确，以及数据库是否存在")
        elif db_type == "mysql":
            import pymysql, socket
            host = config.get("host", "127.0.0.1")
            port = int(config.get("port", 3306))
            user = config.get("username", "")
            password = config.get("password", "")
            database = config.get("database", "")
            # 先用 socket 探测端口是否可达
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(3)
                sock.connect((host, port))
                sock.close()
            except Exception:
                raise Exception(f"MySQL 连接失败：无法访问 {host}:{port}，请确认 MySQL 服务已启动且端口正确")
            try:
                return pymysql.connect(host=host, port=port, user=user, password=password, database=database, autocommit=True)
            except pymysql.OperationalError as e:
                err_msg = str(e) or ""
                if hasattr(e, 'args') and e.args:
                    for arg in e.args:
                        if isinstance(arg, str) and arg.strip():
                            err_msg = arg.strip()
                            break
                if not err_msg:
                    err_msg = f"无法连接到 {host}:{port} 上的数据库 \"{database}\"，请检查用户名、密码是否正确"
                raise Exception(f"MySQL 连接失败：{err_msg}")
        else:
            raise ValueError(f"不支持的数据库类型：{db_type}")

    def api_db_query(self, user):
        """通用数据库查询工具"""
        try:
            data = self.read_json()
            action = data.get("action", "connect")
            config = data.get("config", {})
            db_type = config.get("db_type", "postgres").lower()

            # 如果密码为空或隐藏标记，尝试从 db_connections 表读取真实密码
            pwd = config.get("password", "")
            if not pwd or pwd == "******":
                with get_db_connection() as conn:
                    cur = exec_sql(conn, """
                        SELECT password FROM db_connections
                        WHERE user_id = ? AND db_type = ? AND host = ? AND port = ?
                          AND username = ? AND "database" = ?
                        LIMIT 1
                    """, (user["id"], db_type, config.get("host", ""), int(config.get("port", 0)),
                          config.get("username", ""), config.get("database", "")))
                    row = fetchone(cur)
                    if row and row.get("password"):
                        config["password"] = row["password"]

            if action == "connect":
                conn = self._connect_external_db(config)
                conn.close()
                return self.send_json({"ok": True, "message": "连接成功"})

            conn = self._connect_external_db(config)
            cur = conn.cursor()

            if action == "tables":
                if db_type == "sqlite":
                    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name")
                    tables = [row[0] for row in cur.fetchall()]
                elif db_type == "postgres":
                    cur.execute("""
                        SELECT table_name FROM information_schema.tables
                        WHERE table_schema = 'public' ORDER BY table_name
                    """)
                    tables = [row[0] for row in cur.fetchall()]
                elif db_type == "mysql":
                    cur.execute("SHOW TABLES")
                    tables = [row[0] for row in cur.fetchall()]
                cur.close(); conn.close()
                return self.send_json({"tables": tables})

            if action == "schema":
                table = data.get("table", "")
                if not table:
                    return self.send_json({"error": "缺少表名"}, 400)
                if db_type == "sqlite":
                    cur.execute(f"PRAGMA table_info([{table}])")
                    columns = [{"name": r[1], "type": r[2], "nullable": "NO" if r[3] else "YES", "default": r[4]} for r in cur.fetchall()]
                elif db_type == "postgres":
                    cur.execute("""
                        SELECT column_name, data_type, is_nullable, column_default
                        FROM information_schema.columns
                        WHERE table_name = %s AND table_schema = 'public'
                        ORDER BY ordinal_position
                    """, (table,))
                    columns = [{"name": r[0], "type": r[1], "nullable": r[2], "default": r[3]} for r in cur.fetchall()]
                elif db_type == "mysql":
                    cur.execute("""
                        SELECT column_name, data_type, is_nullable, column_default
                        FROM information_schema.columns
                        WHERE table_name = %s AND table_schema = DATABASE()
                        ORDER BY ordinal_position
                    """, (table,))
                    columns = [{"name": r[0], "type": r[1], "nullable": r[2], "default": r[3]} for r in cur.fetchall()]
                cur.close(); conn.close()
                return self.send_json({"table": table, "columns": columns})

            if action == "query":
                sql = data.get("sql", "").strip()
                if not sql:
                    return self.send_json({"error": "SQL 不能为空"}, 400)
                # 只允许 SELECT / SHOW / DESCRIBE / EXPLAIN
                first_word = sql.split()[0].upper()
                if first_word not in ("SELECT", "SHOW", "DESCRIBE", "DESC", "EXPLAIN"):
                    return self.send_json({"error": "仅支持查询类 SQL（SELECT / SHOW / DESCRIBE / EXPLAIN）"}, 400)
                cur.execute(sql)
                # 获取列名
                col_names = [desc[0] for desc in cur.description] if cur.description else []
                rows = cur.fetchall()
                cur.close(); conn.close()
                return self.send_json({"columns": col_names, "rows": [list(r) for r in rows], "count": len(rows)})

            return self.send_json({"error": "未知操作"}, 400)
        except Exception as e:
            return self.send_json({"error": str(e)}, 400)

    # ========== 数据库连接配置管理 ==========
    def api_get_db_connections(self, user):
        with get_db_connection() as conn:
            cur = exec_sql(conn, "SELECT id, name, db_type, host, port, username, database, sqlite_path, created_at, updated_at FROM db_connections WHERE user_id = ? ORDER BY id", (user["id"],))
            conns = fetchall(cur)
        # 隐藏密码
        for c in conns:
            c["password"] = "******"
        return self.send_json({"connections": conns})

    def api_create_db_connection(self, user):
        try:
            data = self.read_json()
            name = data.get("name", "").strip()
            db_type = data.get("db_type", "").strip()
            host = data.get("host", "").strip()
            port = int(data.get("port", 0))
            username = data.get("username", "").strip()
            password = data.get("password", "")
            database = data.get("database", "").strip()
            sqlite_path = data.get("sqlite_path", "").strip()

            if not name:
                return self.send_json({"error": "连接名称不能为空"}, 400)
            if not db_type:
                return self.send_json({"error": "请选择数据库类型"}, 400)
            if db_type == "sqlite":
                if not sqlite_path:
                    return self.send_json({"error": "SQLite 文件路径不能为空"}, 400)
            else:
                if not host or not username or not database:
                    return self.send_json({"error": "主机、用户名、数据库名为必填项"}, 400)

            from datetime import datetime
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                # 检查同名
                cur = exec_sql(conn, "SELECT id FROM db_connections WHERE user_id = ? AND name = ?", (user["id"], name))
                if fetchone(cur):
                    return self.send_json({"error": f"连接名称 \"{name}\" 已存在"}, 400)
                exec_sql(conn, """
                    INSERT INTO db_connections(user_id, name, db_type, host, port, username, password, database, sqlite_path, created_at, updated_at)
                    VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (user["id"], name, db_type, host, port, username, password, database, sqlite_path, now, now))
                conn.commit()
            return self.send_json({"ok": True, "message": "连接配置已保存"})
        except Exception as e:
            return self.send_json({"error": str(e)}, 400)

    def api_delete_db_connection(self, user, conn_id):
        try:
            conn_id = int(conn_id)
            with get_db_connection() as conn:
                cur = exec_sql(conn, "SELECT id FROM db_connections WHERE id = ? AND user_id = ?", (conn_id, user["id"]))
                if not fetchone(cur):
                    return self.send_json({"error": "连接不存在"}, 404)
                exec_sql(conn, "DELETE FROM db_connections WHERE id = ?", (conn_id,))
                conn.commit()
            return self.send_json({"ok": True, "message": "连接已删除"})
        except Exception as e:
            return self.send_json({"error": str(e)}, 400)

    def api_update_db_connection(self, user, conn_id):
        try:
            conn_id = int(conn_id)
            data = self.read_json()
            name = data.get("name", "").strip()
            db_type = data.get("db_type", "").strip()
            host = data.get("host", "").strip()
            port = int(data.get("port", 0))
            username = data.get("username", "").strip()
            password = data.get("password", "")
            database = data.get("database", "").strip()
            sqlite_path = data.get("sqlite_path", "").strip()

            if not name:
                return self.send_json({"error": "连接名称不能为空"}, 400)
            if not db_type:
                return self.send_json({"error": "请选择数据库类型"}, 400)
            if db_type == "sqlite":
                if not sqlite_path:
                    return self.send_json({"error": "SQLite 文件路径不能为空"}, 400)
            else:
                if not host or not username or not database:
                    return self.send_json({"error": "主机、用户名、数据库名为必填项"}, 400)

            from datetime import datetime
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            with get_db_connection() as conn:
                cur = exec_sql(conn, "SELECT id FROM db_connections WHERE id = ? AND user_id = ?", (conn_id, user["id"]))
                if not fetchone(cur):
                    return self.send_json({"error": "连接不存在"}, 404)
                # 检查同名（排除自身）
                cur = exec_sql(conn, "SELECT id FROM db_connections WHERE user_id = ? AND name = ? AND id != ?", (user["id"], name, conn_id))
                if fetchone(cur):
                    return self.send_json({"error": f"连接名称 \"{name}\" 已存在"}, 400)
                if password:
                    exec_sql(conn, """
                        UPDATE db_connections SET name=?, db_type=?, host=?, port=?, username=?, password=?, database=?, sqlite_path=?, updated_at=?
                        WHERE id=? AND user_id=?
                    """, (name, db_type, host, port, username, password, database, sqlite_path, now, conn_id, user["id"]))
                else:
                    exec_sql(conn, """
                        UPDATE db_connections SET name=?, db_type=?, host=?, port=?, username=?, database=?, sqlite_path=?, updated_at=?
                        WHERE id=? AND user_id=?
                    """, (name, db_type, host, port, username, database, sqlite_path, now, conn_id, user["id"]))
                conn.commit()
            return self.send_json({"ok": True, "message": "连接已更新"})
        except Exception as e:
            return self.send_json({"error": str(e)}, 400)

def main():
    init_db()
    server = ThreadingHTTPServer((APP_HOST, APP_PORT), Handler)
    print(f"小肖的自用工具已启动：http://127.0.0.1:{APP_PORT}")
    server.serve_forever()

if __name__ == "__main__":
    main()