# -*- coding: utf-8 -*-
"""Excel 导入导出业务逻辑"""
import csv
import io
from datetime import datetime
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _safe_get(row: dict, key: str, default=""):
    """安全获取字典值，None 转为默认值。"""
    try:
        val = row.get(key)
        return val if val is not None else default
    except (KeyError, IndexError, TypeError):
        return default


def export_csv(rows: List[dict]) -> bytes:
    """导出 CSV（带 BOM 头，Excel 友好）。"""
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(
        ["ID", "日期", "类型", "项目", "金额", "消费分类", "账户", "备注", "创建时间", "更新时间"]
    )
    for r in rows:
        writer.writerow(
            [
                r["id"],
                r["record_date"],
                "收入" if r["type"] == "income" else "支出",
                r["category"],
                r["amount"],
                r.get("sub_category", "") or "",
                r.get("account", "") or "",
                r.get("note", "") or "",
                r["created_at"],
                r["updated_at"],
            ]
        )
    return ("\ufeff" + out.getvalue()).encode("utf-8")


def export_excel(rows: List[dict], output: io.BytesIO = None) -> bytes:
    """导出带样式的 Excel。若传入 output buffer 则写入并返回其内容，否则返回 bytes。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "记账记录"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")
    headers = ["ID", "日期", "类型", "项目", "金额", "消费分类", "账户", "备注", "创建时间", "更新时间"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    for row_num, r in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=_safe_get(r, "id", 0))
        ws.cell(row=row_num, column=2, value=_safe_get(r, "record_date"))
        ws.cell(row=row_num, column=3, value="收入" if _safe_get(r, "type") == "income" else "支出")
        ws.cell(row=row_num, column=4, value=_safe_get(r, "category"))
        ws.cell(row=row_num, column=5, value=_safe_get(r, "amount", 0))
        ws.cell(row=row_num, column=6, value=_safe_get(r, "sub_category"))
        ws.cell(row=row_num, column=7, value=_safe_get(r, "account"))
        ws.cell(row=row_num, column=8, value=_safe_get(r, "note"))
        ws.cell(row=row_num, column=9, value=_safe_get(r, "created_at"))
        ws.cell(row=row_num, column=10, value=_safe_get(r, "updated_at"))

    if output is None:
        output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def download_template() -> bytes:
    """生成导入模板（含"导入数据"和"字段说明"两页）。"""
    wb = Workbook()

    # ====== 第1页：填写表格 ======
    ws1 = wb.active
    ws1.title = "导入数据"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    headers = ["日期", "类型", "项目", "金额", "消费分类", "账户", "备注", "星期几", "是否取消"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        if header in ["日期", "类型", "项目", "金额"]:
            ws1.cell(row=2, column=col, value="* 必填")
            ws1.cell(row=2, column=col).fill = required_fill
            ws1.cell(row=2, column=col).alignment = Alignment(horizontal="center")
        else:
            ws1.cell(row=2, column=col, value="选填")
            ws1.cell(row=2, column=col).alignment = Alignment(horizontal="center")

    sample_data = [
        ["2025/06/17", "支出", "午餐", 6.85, "餐饮", "微信", "", "星期四", ""],
        ["2025/06/17", "支出", "地铁", 4, "交通", "支付宝", "上班通勤", "星期四", ""],
        ["2025/06/17", "收入", "工资", 15000, "收入", "银行卡", "2025年6月工资", "星期五", ""],
    ]
    for row_idx, row_data in enumerate(sample_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=row_idx, column=col_idx, value=value)

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 28
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 18
    ws1.column_dimensions["F"].width = 12
    ws1.column_dimensions["G"].width = 25
    ws1.freeze_panes = "A3"

    # ====== 第2页：字段说明 ======
    ws2 = wb.create_sheet(title="字段说明")

    title_font = Font(bold=True, size=14)
    ws2.cell(row=1, column=1, value="记账导入模板 - 字段说明")
    ws2.cell(row=1, column=1).font = title_font
    ws2.merge_cells("A1:D1")

    field_headers = ["字段名称", "是否必填", "格式要求", "示例"]
    for col, header in enumerate(field_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    field_descriptions = [
        ["日期", "必填", "支持多种日期格式\nYYYY/MM/DD\nYYYY-MM-DD\nYYYY年MM月DD日",
         "2025/06/17\n2025-06-17\n2025年6月17日"],
        ["类型", "必填", "收入 / 支出", "收入\n支出"],
        ["项目", "必填", "消费项目或收入来源", "午餐\n工资\n地铁"],
        ["金额", "必填", "数字格式\n负数自动识别为支出", "6.85\n-4.00\n15000"],
        ["消费分类", "选填", "更详细的分类标签", "餐饮\n收入\n交通"],
        ["账户", "选填", "支付或收款账户", "微信\n支付宝\n银行卡"],
        ["备注", "选填", "额外说明文字", "和同事聚餐"],
    ]

    for row_idx, row_data in enumerate(field_descriptions, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 2 and "必填" in str(value):
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 35
    ws2.column_dimensions["D"].width = 30
    for row in range(4, 11):
        ws2.row_dimensions[row].height = 45

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def import_excel(user_id: int, file_data: bytes) -> dict:
    """导入 Excel 文件。自动识别表头行，并读取其下方所有非空数据行。"""
    excel_file = io.BytesIO(file_data)
    wb = load_workbook(excel_file, data_only=True)
    ws = wb.active

    # 自动定位表头行：包含"日期"和"项目"列
    header_row = 1
    headers = {}
    for row in range(1, min(ws.max_row, 10) + 1):
        row_headers = {}
        for col in range(1, ws.max_column + 1):
            val = str(ws.cell(row=row, column=col).value or "").strip()
            if val:
                row_headers[val] = col
        if "日期" in row_headers and "项目" in row_headers:
            header_row = row
            headers = row_headers
            break

    required = ["日期", "类型", "项目", "金额"]
    for r in required:
        if r not in headers:
            raise ValueError(f"缺少必须列：{r}")

    success_count = 0
    errors = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    type_map = {"收入": "income", "支出": "expense"}

    from app.database import exec_sql, get_db

    with get_db() as conn:
        for row_num in range(header_row + 1, ws.max_row + 1):
            try:
                record_date = str(ws.cell(row=row_num, column=headers["日期"]).value or "").strip()
                record_type = str(ws.cell(row=row_num, column=headers["类型"]).value or "").strip()
                category = str(ws.cell(row=row_num, column=headers["项目"]).value or "").strip()
                amount_val = ws.cell(row=row_num, column=headers["金额"]).value

                sub_category = ""
                if headers.get("消费分类"):
                    sub_category = str(ws.cell(row=row_num, column=headers["消费分类"]).value or "").strip()
                note = ""
                if headers.get("备注"):
                    note = str(ws.cell(row=row_num, column=headers["备注"]).value or "").strip()
                account = ""
                if headers.get("账户"):
                    account = str(ws.cell(row=row_num, column=headers["账户"]).value or "").strip()

                # 跳过空行或示例/标记行（无日期且无项目）
                if not record_date and not category:
                    continue

                # 跳过示例占位行（如"* 必填"、"选填"等）
                if record_date in ["* 必填", "选填"] or category in ["* 必填", "选填"]:
                    continue

                # 解析日期
                parsed_date = None
                if isinstance(record_date, datetime):
                    parsed_date = record_date
                else:
                    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S"):
                        try:
                            parsed_date = datetime.strptime(str(record_date), fmt)
                            break
                        except ValueError:
                            continue
                if not parsed_date:
                    errors.append(f"第 {row_num} 行：日期格式无效")
                    continue
                record_date = parsed_date.strftime("%Y-%m-%d")

                if record_type not in type_map:
                    errors.append(f"第 {row_num} 行：类型只能是收入或支出")
                    continue
                record_type = type_map[record_type]

                try:
                    amount = round(float(amount_val or 0), 2)
                except (ValueError, TypeError):
                    errors.append(f"第 {row_num} 行：金额格式无效")
                    continue

                if amount < 0:
                    amount = abs(amount)
                    if record_type == "income":
                        record_type = "expense"

                if not category:
                    errors.append(f"第 {row_num} 行：项目不能为空")
                    continue

                exec_sql(
                    conn,
                    """INSERT INTO records
                       (user_id, record_date, type, category, sub_category, amount, account, note, created_at, updated_at)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                    (user_id, record_date, record_type, category, sub_category, amount, account, note, now, now),
                )
                success_count += 1
            except Exception as e:
                errors.append(f"第 {row_num} 行：{str(e)}")

    return {"success": success_count, "errors": errors}
